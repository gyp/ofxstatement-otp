from openpyxl import load_workbook
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
import logging
import warnings
from typing import Optional

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import Statement, StatementLine

from ofxstatement_otp.transaction_types import transaction_type

logger = logging.getLogger("OTP")

TRANSACTIONS_SHEET_NAME = "Tranzakciók"

# Labels that mark the metadata cells in the preamble (column A, value in
# column B). The post-2025 OTP netbank export prefixes the transaction table
# with a few rows of statement metadata.
LABEL_START_DATE = "Lekérdezés kezdete"
LABEL_END_DATE = "Lekérdezés vége"

# Column A header that marks the start of the transaction table. The same word
# also appears as a metadata label in the preamble, so the header row is
# identified by also requiring the column B header (see _find_header_row).
HEADER_ACCOUNT_NO = "Számlaszám"
HEADER_PARTNER_ACCOUNT = "Ellenoldali számlaszám"

DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
DATE_FORMAT = "%Y-%m-%d"


@dataclass
class Transaction:
    account_no: str
    partner_account: str
    partner_name: str
    description: str
    memo: str
    category: str
    bank_txn_id: str
    transaction_date: Optional[datetime]
    booking_date: datetime
    amount: Decimal
    currency: str


class OtpPlugin(Plugin):
    """OTP Bank, post-2026-June netbank export (XLSX).

    The export bundles every account into a single file; set the ``account``
    setting to the (partial) account number to emit a statement for just one
    account.
    """

    def get_parser(self, filename):
        parser = OtpXlsxParser(filename, self.settings)
        parser.statement.bank_id = self.settings.get("BIC", "OTPVHUHB")
        return parser


def _to_datetime(value, fmt: str) -> datetime:
    """Accept either a native datetime cell or a string in ``fmt``."""
    if isinstance(value, datetime):
        return value
    return datetime.strptime(value, fmt)


def _load_workbook(filename):
    """Load an xlsx, silencing openpyxl's harmless "no default style" warning.

    OTP exports ship a stylesheet without a named default style, which makes
    openpyxl warn on every load. The warning is noise for our read-only use.
    """
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style",
            category=UserWarning,
        )
        return load_workbook(filename)


class OtpXlsxParser(StatementParser):

    def __init__(self, filename, settings=None):
        self.filename = filename
        self.settings = settings or {}
        self.account_filter = self.settings.get("account")
        self.workbook = _load_workbook(self.filename)
        self.sheet = self.workbook[TRANSACTIONS_SHEET_NAME]
        self.header_row = self._find_header_row()

        self.statement = Statement()
        self.statement.account_id = self._get_account_id()
        self.statement.currency = self._get_currency()
        self.statement.start_balance = self._get_start_balance()
        self.statement.start_date = self._get_start_date()
        self.statement.end_balance = self._get_end_balance()
        self.statement.end_date = self._get_end_date()
        logger.debug(self.statement)

    def parse(self):
        return super(OtpXlsxParser, self).parse()

    def split_records(self):
        return self._get_transactions()

    def parse_record(self, record: Transaction):
        logger.debug(record)
        stat_line = StatementLine(
            record.bank_txn_id,
            record.booking_date,
            record.memo,
            Decimal(str(record.amount)),
        )
        stat_line.date_user = record.transaction_date
        stat_line.payee = record.partner_name
        stat_line.trntype = self._get_transaction_type(record)
        logger.debug(stat_line)
        return stat_line

    def _find_header_row(self) -> int:
        """Locate the transaction table header row.

        ``Számlaszám`` appears both as a metadata label and as the table's
        first column header, so the header is disambiguated by also requiring
        the partner-account header in column B.
        """
        for row in range(1, self.sheet.max_row + 1):
            a = self.sheet.cell(row=row, column=1).value
            b = self.sheet.cell(row=row, column=2).value
            if a == HEADER_ACCOUNT_NO and b == HEADER_PARTNER_ACCOUNT:
                return row
        raise ValueError(
            "Could not find the transaction table header (expected a row with "
            f"'{HEADER_ACCOUNT_NO}' / '{HEADER_PARTNER_ACCOUNT}') in sheet "
            f"'{TRANSACTIONS_SHEET_NAME}'"
        )

    def _included(self, account_no) -> bool:
        if not self.account_filter:
            return True
        return self.account_filter.lower() in str(account_no).lower()

    def _get_transactions(self):
        for row in range(self.header_row + 1, self.sheet.max_row + 1):
            # skip hidden rows -- some netbank filters are applied as such
            if self.sheet.row_dimensions[row].hidden:
                logger.debug("Skipping hidden row: %s", row)
                continue

            cells = [self.sheet.cell(row=row, column=col).value for col in range(1, 12)]
            (
                account_no,
                partner_account,
                partner_name,
                description,
                memo,
                category,
                bank_txn_id,
                transaction_date,
                booking_date,
                amount,
                currency,
            ) = cells

            # stop at the first row without an account number
            if not account_no:
                break

            # skip rows without a booking date (e.g. pending items)
            if not booking_date:
                logger.debug("Skipping incomplete row: %s", row)
                continue

            # only emit the configured account, if filtering is enabled
            if not self._included(account_no):
                continue

            yield Transaction(
                account_no=account_no,
                partner_account=partner_account,
                partner_name=partner_name,
                description=description,
                memo=memo,
                category=category,
                bank_txn_id=bank_txn_id,
                # The row is booked (booking date present, checked above); the
                # transaction datetime may still be blank, so guard it.
                transaction_date=(
                    _to_datetime(transaction_date, DATETIME_FORMAT)
                    if transaction_date
                    else None
                ),
                booking_date=_to_datetime(booking_date, DATE_FORMAT),
                amount=amount,
                currency=currency,
            )

    def _get_account_id(self) -> Optional[str]:
        # When filtering, report the filtered account; otherwise fall back to
        # the first account number that appears in the data.
        for row in range(self.header_row + 1, self.sheet.max_row + 1):
            account_no = self.sheet.cell(row=row, column=1).value
            if not account_no:
                break
            if self._included(account_no):
                return str(account_no)
        if self.account_filter:
            logger.warning(
                "No transactions matched account filter %r; "
                "leaving account_id unset",
                self.account_filter,
            )
        return None

    def _get_currency(self) -> str:
        # Read the currency of the first matching transaction; default to HUF.
        for row in range(self.header_row + 1, self.sheet.max_row + 1):
            account_no = self.sheet.cell(row=row, column=1).value
            if not account_no:
                break
            if self._included(account_no):
                currency = self.sheet.cell(row=row, column=11).value
                if currency:
                    return str(currency)
        return "HUF"

    def _get_preamble_value(self, label: str):
        """Return the column-B value of a labelled metadata row (rows above
        the transaction table)."""
        for row in range(1, self.header_row):
            if self.sheet.cell(row=row, column=1).value == label:
                return self.sheet.cell(row=row, column=2).value
        return None

    def _get_start_date(self) -> Optional[datetime]:
        value = self._get_preamble_value(LABEL_START_DATE)
        return _to_datetime(value, DATE_FORMAT) if value else None

    def _get_end_date(self) -> Optional[datetime]:
        value = self._get_preamble_value(LABEL_END_DATE)
        return _to_datetime(value, DATE_FORMAT) if value else None

    def _get_start_balance(self):
        return None

    def _get_end_balance(self):
        return None

    def _get_transaction_type(self, transaction: Transaction) -> str:
        return transaction_type(transaction.description)
