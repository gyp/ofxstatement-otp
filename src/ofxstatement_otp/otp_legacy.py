"""OTP netbank XLSX export, pre-2026-June format.

This is the format the ``otp`` plugin read before OTP changed the export in
June 2026 (data from row 2, columns A-L, account id from cell D8 and the start
date from cell B2). It is kept as ``otp_legacy`` for converting older exports;
new exports use the ``otp`` plugin.
"""

from openpyxl import load_workbook
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
import logging

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import Statement, StatementLine, generate_transaction_id

from ofxstatement_otp.transaction_types import transaction_type

logger = logging.getLogger("OTP")

TRANSACTIONS_SHEET_NAME = "Tranzakciók"


@dataclass
class Transaction:
    transaction_date: datetime
    booking_date: datetime
    description: str
    in_or_out: str
    partner_name: str
    partner_account: str
    otp_generated_category: str
    memo: str
    account_name: str
    account_no: str
    amount: Decimal
    currency: str


class OtpLegacyPlugin(Plugin):
    """OTP Bank, pre-2026-June netbank export (XLSX)."""

    def get_parser(self, filename):
        parser = OtpLegacyXlsxParser(filename)
        parser.statement.bank_id = self.settings.get("BIC", "OTPVHUHB")
        return parser


class OtpLegacyXlsxParser(StatementParser):

    def split_records(self):
        return self._get_transactions()

    def __init__(self, filename):
        self.filename = filename
        self.statement = Statement()
        self.statement.account_id = self._get_account_id()
        self.statement.currency = self._get_currency()
        self.statement.start_balance = self._get_start_balance()
        self.statement.start_date = self._get_start_date()
        self.statement.end_balance = self._get_end_balance()
        self.statement.end_date = self._get_end_date()
        logger.debug(self.statement)

    def parse(self):
        return super(OtpLegacyXlsxParser, self).parse()

    def parse_record(self, record: Transaction):
        logger.debug(record)
        stat_line = StatementLine(
            None, record.booking_date, record.memo, Decimal(record.amount)
        )
        logger.debug(stat_line)
        stat_line.id = generate_transaction_id(stat_line)
        stat_line.date_user = record.transaction_date
        stat_line.payee = record.partner_name
        stat_line.trntype = self._get_transaction_type(record)
        logger.debug(stat_line)
        return stat_line

    def _get_account_id(self):
        # FIXME: there's no single field for this in this format
        wb = load_workbook(self.filename)
        return wb[TRANSACTIONS_SHEET_NAME]["D8"].value

    def _get_currency(self):
        # TODO: support non-HUF accounts
        return "HUF"

    def _get_transactions(self):
        wb = load_workbook(self.filename)
        starting_column = "A"
        ending_column = "L"
        starting_row = 2
        offset = 0

        while True:
            data = wb[TRANSACTIONS_SHEET_NAME][
                "{}{}:{}{}".format(
                    starting_column,
                    starting_row + offset,
                    ending_column,
                    starting_row + offset,
                )
            ]

            cells = [*map(lambda x: x.value, data[0])]
            logger.debug(cells)
            (
                transaction_date,
                booking_date,
                description,
                in_or_out,
                partner_name,
                partner_account,
                otp_generated_category,
                memo,
                account_name,
                account_no,
                amount,
                currency,
            ) = cells

            # skip hidden rows -- some filters are applied as such
            if wb.active.row_dimensions[starting_row + offset].hidden:
                logger.debug("Skipping hidden row: " + str(starting_row + offset))
                offset += 1
                continue

            # stop when we reach the end of the file
            if not transaction_date:
                break

            # skip lines without parseable dates
            if not booking_date:
                logger.debug("Skipping incomplete row: " + str(starting_row + offset))
                offset += 1
                continue

            offset += 1
            yield Transaction(
                transaction_date=datetime.strptime(
                    transaction_date, "%Y-%m-%d %H:%M:%S"
                ),
                booking_date=datetime.strptime(booking_date, "%Y-%m-%d"),
                description=description,
                in_or_out=in_or_out,
                partner_name=partner_name,
                partner_account=partner_account,
                otp_generated_category=otp_generated_category,
                memo=memo,
                account_name=account_name,
                account_no=account_no,
                amount=amount,
                currency=currency,
            )

    def _get_transaction_type(self, transaction: Transaction) -> str:
        return transaction_type(transaction.description)

    def _get_start_balance(self):
        return None

    def _get_end_balance(self):
        return None

    def _get_start_date(self):
        wb = load_workbook(self.filename)
        date = wb[TRANSACTIONS_SHEET_NAME]["B2"].value
        return datetime.strptime(date, "%Y-%m-%d")

    def _get_end_date(self):
        return None
