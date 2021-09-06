from openpyxl import load_workbook
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
import logging

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import (Statement, StatementLine,
                                    generate_transaction_id)

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger('OTP')

TRANSACTIONS_SHEET_NAME = 'Tranzakciók'

@dataclass
class Transaction:
    booking_date: datetime
    transaction_date: datetime
    description: str
    credit: Decimal
    debit: Decimal
    description_details: str
    notes: str


class OtpPlugin(Plugin):

    def get_parser(self, filename):
        parser = OtpXlsxParser(filename)
        parser.statement.bank_id = self.settings.get('BIC', 'IntesaSP')
        return parser


class OtpXlsxParser(StatementParser):

    def split_records(self):
        return self._get_transactions()

    def __init__(self, filename):
        self.filename = filename
        self.statement = Statement()
        self.statement.account_id = self._get_account_id()
        self.statement.currency = self._get_currency()
        self.statement.start_balance = self._get_start_balance()
        self.statement.start_date = self._get_start_date()
        self.statement.end_balance = self._get_end_balance()
        self.statement.end_date = self._get_end_date()
        logging.debug(self.statement)

    def parse(self):
        return super(OtpXlsxParser, self).parse()

    def parse_record(self, record: Transaction):
        stat_line = StatementLine(None,
                                  record.booking_date,
                                  record.transaction_date,
                                  Decimal(record.credit) if record.credit else
                                  Decimal(record.debit))
        stat_line.id = generate_transaction_id(stat_line)
        stat_line.date_user = record.data_valuta
        stat_line.trntype = OtpXlsxParser._get_transaction_type(record)
        logging.debug(stat_line)
        return stat_line

    def _get_account_id(self):
        # FIXME: there's no single field for this in this format
        wb = load_workbook(self.filename)
        return wb[TRANSACTIONS_SHEET_NAME]['D8'].value

    def _get_currency(self):
        # TODO: support non-HUF accounts
        return 'HUF'

    def _get_transactions(self) -> list[Transaction]:
        wb = load_workbook(self.filename)
        starting_column = 'A'
        ending_column = 'L'
        starting_row = 1
        offset = 0

        while True:
            data = wb[TRANSACTIONS_SHEET_NAME][
                        '{}{}:{}{}'.format(starting_column,
                                           starting_row + offset,
                                           ending_column,
                                           starting_row + offset
                                           )]

            values = [*map(lambda x: x.value, data[0])]
            logging.debug(values)
            if not values[0]:
                break
            else:
                yield Transaction(*values)
                offset += 1

    def _get_transaction_type(self, transaction: Transaction) -> str:
        # FIXME: update this to the Hungarian transaction types
        trans_map = {'Pagamento pos': 'POS',
                     'Pagamento effettuato su pos estero': 'POS',
                     'Accredito beu con contabile': 'XFER',
                     'Canone mensile base e servizi aggiuntivi': 'SRVCHG',
                     'Prelievo carta debito su banche del gruppo': 'CASH',
                     'Prelievo carta debito su banche italia/sepa': 'CASH',
                     'Comm.prelievo carta debito italia/sepa': 'SRVCHG',
                     'Commiss. su beu internet banking': 'SRVCHG',
                     'Pagamento telefono': 'PAYMENT',
                     'Pagamento mav via internet banking': 'PAYMENT',
                     'Pagamento bolletta cbill': 'PAYMENT',
                     'Beu tramite internet banking': 'PAYMENT',
                     'Commissione bolletta cbill': 'SRVCHG',
                     'Storno pagamento pos': 'POS',
                     'Storno pagamento pos estero': 'POS',
                     'Versamento contanti su sportello automatico': 'ATM',
                     'Canone annuo o-key sms': 'SRVCHG'
                     }
        return trans_map[transaction.description]

    def _get_start_balance(self):
        # FIXME: this export does not have balance information 
        wb = load_workbook(self.filename)
        return Decimal(wb[TRANSACTIONS_SHEET_NAME]['E11'].value)

    def _get_end_balance(self):
        # FIXME: this export does not have balance information
        wb = load_workbook(self.filename)
        return Decimal(wb[TRANSACTIONS_SHEET_NAME]['E12'].value)

    def _get_start_date(self):
        wb = load_workbook(self.filename)
        date = wb[TRANSACTIONS_SHEET_NAME]['B2'].value
        return datetime.strptime(date, '%d-%m-%Y')

    def _get_end_date(self):
        # FIXME: there's no explicit field for this, we need to find the last row for this
        wb = load_workbook(self.filename)
        date = wb[TRANSACTIONS_SHEET_NAME]['D12'].value
        return datetime.strptime(date, '%d.%m.%Y')
